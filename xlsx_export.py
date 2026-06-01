"""Build a classified-SERP .xlsx matching the boss's template."""
from __future__ import annotations

from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import DoughnutChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import CATEGORIES, BY_KEY

# Body fills derived from the central taxonomy: key -> (bg, font, bold)
BODY_FILL = {c.key: (c.fill, c.font, c.bold) for c in CATEGORIES}

# Legend rows: (display label, definition) in taxonomy order
LEGEND_ROWS = [(c.label, c.definition) for c in CATEGORIES]

# Summary block lists every category key in taxonomy order
SUMMARY_ROWS = [c.key for c in CATEGORIES]

HEADER_NAVY = "1F3864"
HEADER_GREY = "D8D8D8"

# ── Layout anchors — computed so the sheet never collides as the taxonomy grows ──
LEGEND_HEADER_ROW = 5
LEGEND_START = 6
_LEGEND_END = LEGEND_START + len(CATEGORIES) - 1     # 11 cats -> row 16
KEYWORD_HEADER_ROW = _LEGEND_END + 2                 # gap of 1 -> row 18
KEYWORD_START = KEYWORD_HEADER_ROW + 1               # row 19
NUM_KEYWORDS = 26
_KEYWORD_END = KEYWORD_START + NUM_KEYWORDS - 1      # row 44
SUMMARY_HEADER_ROW = _KEYWORD_END + 4                # gap of 3 -> row 48
SUMMARY_START = SUMMARY_HEADER_ROW + 1               # row 49
_SUMMARY_END = SUMMARY_START + len(CATEGORIES) - 1   # row 59
TOTAL_ROW = _SUMMARY_END + 1                         # row 60


def _fill(rgb: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=rgb)


def _category_style(cell, category: str) -> None:
    if category not in BODY_FILL:
        return
    bg, fg, bold = BODY_FILL[category]
    cell.fill = _fill(bg)
    cell.font = Font(color=fg, bold=bold)


def _format_date(run_date: str) -> str:
    try:
        dt = datetime.strptime(run_date, "%Y-%m-%d")
    except ValueError:
        return run_date
    # "Date: May 18, 2026" — strip leading zero on day for non-Windows portability
    return f"Date: {dt.strftime('%B')} {dt.day}, {dt.year}"


def build_workbook(run: dict) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "rankings"

    # Row 2: date
    c = ws.cell(row=2, column=1, value=_format_date(run["run_date"]))
    c.font = Font(bold=True)

    # Row 3: VPN line
    c = ws.cell(row=3, column=1, value="VPN mobile")
    c.font = Font(bold=True)

    # Legend header
    h1 = ws.cell(row=LEGEND_HEADER_ROW, column=1, value="Label")
    h2 = ws.cell(row=LEGEND_HEADER_ROW, column=2, value="Meaning")
    for cell in (h1, h2):
        cell.fill = _fill(HEADER_NAVY)
        cell.font = Font(color="FFFFFF", bold=True)

    # Legend body — one row per category, styled in its own colour
    for i, ((label, meaning), cat_key) in enumerate(zip(LEGEND_ROWS, SUMMARY_ROWS)):
        r = LEGEND_START + i
        label_cell = ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=meaning)
        _category_style(label_cell, cat_key)

    # Keyword block column headers (Keyword + 1..10)
    hdr = ws.cell(row=KEYWORD_HEADER_ROW, column=1, value="Keyword")
    hdr.fill = _fill(HEADER_NAVY)
    hdr.font = Font(color="FFFFFF", bold=True)
    for i in range(10):
        c = ws.cell(row=KEYWORD_HEADER_ROW, column=2 + i, value=i + 1)
        c.fill = _fill(HEADER_NAVY)
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center")

    # 26 keyword rows
    counts = {k: 0 for k in SUMMARY_ROWS}
    for row_idx, kw in enumerate(run["keywords"]):
        r = KEYWORD_START + row_idx
        ws.cell(row=r, column=1, value=kw["keyword"])
        positions_by_rank = {p["rank"]: p for p in kw.get("positions", [])}
        for rank in range(1, 11):
            p = positions_by_rank.get(rank)
            cell = ws.cell(row=r, column=1 + rank)
            if not p:
                continue
            cell.value = p.get("short_label") or p.get("domain", "")
            cat = p.get("category", "")
            _category_style(cell, cat)
            if cat in counts:
                counts[cat] += 1

    # Summary header
    for i, val in enumerate(["Category", "MARKET SHARE %", "Count"]):
        c = ws.cell(row=SUMMARY_HEADER_ROW, column=1 + i, value=val)
        c.fill = _fill(HEADER_GREY)
        c.font = Font(bold=True)

    # Per-category counts — label cell coloured in its own category colour
    for i, cat in enumerate(SUMMARY_ROWS):
        r = SUMMARY_START + i
        label = ws.cell(row=r, column=1, value=BY_KEY[cat].label)
        _category_style(label, cat)
        ws.cell(row=r, column=2, value=f"=C{r}/C{TOTAL_ROW}").number_format = "0.00%"
        ws.cell(row=r, column=3, value=counts[cat])

    # TOTAL row
    total_label = ws.cell(row=TOTAL_ROW, column=1, value="TOTAL")
    total_label.font = Font(bold=True)
    bsum = ws.cell(row=TOTAL_ROW, column=2,
                   value=f"=SUM(B{SUMMARY_START}:B{_SUMMARY_END})")
    bsum.font = Font(bold=True)
    bsum.number_format = "0.00%"
    csum = ws.cell(row=TOTAL_ROW, column=3,
                   value=f"=SUM(C{SUMMARY_START}:C{_SUMMARY_END})")
    csum.font = Font(bold=True)

    # Column widths — roughly match template's main columns
    widths = {1: 44.0, 2: 22.0, 3: 22.0, 4: 22.0, 5: 22.0, 6: 22.0,
              7: 22.0, 8: 22.0, 9: 22.0, 10: 22.0, 11: 22.0}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Doughnut chart of category market share, anchored to the right of the
    # summary block. Slice colours match the taxonomy palette so the chart reads
    # against the colour key.
    chart = DoughnutChart()
    chart.holeSize = 50
    data_ref = Reference(ws, min_col=2, min_row=SUMMARY_START, max_row=_SUMMARY_END)
    cats_ref = Reference(ws, min_col=1, min_row=SUMMARY_START, max_row=_SUMMARY_END)
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    series = chart.series[0]
    series.dPt = []
    for i, cat in enumerate(SUMMARY_ROWS):
        bg = BODY_FILL[cat][0]
        gp = GraphicalProperties(solidFill=bg)
        gp.line = None
        series.dPt.append(DataPoint(idx=i, spPr=gp))
    chart.width = 15
    chart.height = 7.5
    ws.add_chart(chart, f"E{SUMMARY_HEADER_ROW}")

    # Match template print setup
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    return wb
